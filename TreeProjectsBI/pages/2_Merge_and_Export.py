# -*- coding: utf-8 -*-
from __future__ import annotations

import io
import re
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from style_pack import inject_base_css, apply_plotly_theme, hero_header, glass_container
from utils_he import (
    TARGET_COLS,
    map_col,
    detect_header_row,
    clean_text,
)

# ===================== UI / DESIGN =====================
inject_base_css(bg_main="assets/bg_main.jpg", bg_sidebar="assets/bg_sidebar.jpg")
apply_plotly_theme()
hero_header("🧩 Merge & Export", "מיזוג, פיענוח והמרות — גרסת BI משודרגת.")

st.title("🌳 forestry_and_trees_report2024 — דוחות כריתה אחוד")
st.caption(
    "טען/י: 1) קובץ הדוחות המאוחד (ללא רשימות), "
    "2) קובץ 'רשימת ערים לפי קודים', "
    "3) קובץ 'רשימת עצים לפי קודים'."
)

with glass_container():
    main_file = st.file_uploader(
        "📁 קובץ דוחות כריתה אחוד (ללא רשימות עצים/יישובים)",
        type=["xlsx"],
        key="main",
    )
    city_file = st.file_uploader(
        "🏙️ קובץ 'רשימת ערים לפי קודים'",
        type=["xlsx"],
        key="cities",
    )
    tree_file = st.file_uploader(
        "🌳 קובץ 'רשימת עצים לפי קודים'",
        type=["xlsx"],
        key="trees",
    )

run_btn = st.button("🚀 הרץ מיזוג והמרות")


# ===================== HELPERS =====================

def get_series(df: pd.DataFrame, col_name: str) -> pd.Series:
    """אם יש עמודות כפולות באותו השם — מאחד עם bfill ולוקח עמודה אחת."""
    mask = (df.columns == col_name)
    cols = df.loc[:, mask]
    if cols.shape[1] == 0:
        return pd.Series([pd.NA] * len(df), index=df.index)
    if cols.shape[1] == 1:
        return cols.iloc[:, 0]
    return cols.bfill(axis=1).iloc[:, 0]


def safe_to_datetime_series(s: pd.Series) -> pd.Series:
    """המרת תאריכים לעמידה לשגיאות כולל מספרים סריאליים של אקסל."""
    dt = pd.to_datetime(s, errors="coerce", infer_datetime_format=True, dayfirst=True)
    need = dt.isna()
    if need.any():
        numeric = pd.to_numeric(s[need], errors="coerce")
        good = numeric[numeric.notna()]
        if not good.empty:
            dt.loc[good.index] = pd.to_datetime(
                good,
                unit="D",
                origin="1899-12-30",
                errors="coerce",
            )
    return dt


# ---------- פיענוח קודים: פעולה + סיבה ----------

ACTION_MAP = {1: "כריתה", 2: "העתקה"}

REASON_MAP = {
    1: "אחר",
    2: "בטיחות",
    3: "מחלת עץ",
    4: "סכנה בריאותית",
    5: "בנייה",
    6: "הכשרה חקלאית",
    7: "עץ מת",
    8: "דילול יער",
    9: "קריאה",
    10: "סניטציה",
}


def _to_int_or_nan(val):
    if pd.isna(val):
        return np.nan
    try:
        f = float(str(val).strip())
        i = int(f)
        return i if f == i else np.nan
    except Exception:
        return np.nan


def decode_action_reason(df: pd.DataFrame) -> pd.DataFrame:
    """
    מייצר:
    - פעולה_מפוענחת
    - פעולה_מפוענחת (2)
    - סיבה_מפוענחת
    רק כאשר מדובר בקוד מספרי; טקסט חופשי נשאר כמו שהוא.
    """
    out = df.copy()

    # פעולה ראשית
    if "פעולה" in out.columns:
        act = out["פעולה"].map(_to_int_or_nan)
        out["פעולה_מפוענחת"] = np.where(
            act.notna(),
            act.map(ACTION_MAP).fillna(out["פעולה"].astype(str)),
            out["פעולה"].astype(str),
        )

    # פעולה שנייה
    if "פעולה (2)" in out.columns:
        act2 = out["פעולה (2)"].map(_to_int_or_nan)
        out["פעולה_מפוענחת (2)"] = np.where(
            act2.notna(),
            act2.map(ACTION_MAP).fillna(out["פעולה (2)"].astype(str)),
            out["פעולה (2)"].astype(str),
        )

    # סיבה
    if "סיבה" in out.columns:
        rea = out["סיבה"].map(_to_int_or_nan)
        out["סיבה_מפוענחת"] = np.where(
            rea.notna(),
            rea.map(REASON_MAP).fillna(out["סיבה"].astype(str)),
            out["סיבה"].astype(str),
        )
    elif "סיבה  מילולית" in out.columns:
        out["סיבה_מפוענחת"] = out["סיבה  מילולית"].astype(str)

    return out


def reorder_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    מסדר את העמודות כך ש:
    - 'פעולה' ליד 'פעולה_מפוענחת'
    - 'פעולה (2)' ליד 'פעולה_מפוענחת (2)'
    - 'סיבה' / 'סיבה  מילולית' ליד 'סיבה_מפוענחת'
    ושאר העמודות נשארות באותו סדר יחסי.
    """
    cols = list(df.columns)
    new_order: list[str] = []

    # פעולה 1
    if "פעולה" in cols:
        new_order.append("פעולה")
    if "פעולה_מפוענחת" in cols:
        new_order.append("פעולה_מפוענחת")

    # פעולה 2
    if "פעולה (2)" in cols:
        new_order.append("פעולה (2)")
    if "פעולה_מפוענחת (2)" in cols:
        new_order.append("פעולה_מפוענחת (2)")

    # סיבות
    if "סיבה" in cols:
        new_order.append("סיבה")
    if "סיבה_מפוענחת" in cols:
        new_order.append("סיבה_מפוענחת")
    if "סיבה  מילולית" in cols:
        new_order.append("סיבה  מילולית")

    # שאר העמודות בסדר המקורי
    for c in cols:
        if c not in new_order:
            new_order.append(c)

    return df[new_order].copy()


# ---------- לוקאפ ערים / עצים מקבצי המפתח ----------

def _load_city_lut(city_file) -> dict[int, str]:
    """
    קורא את קובץ 'רשימת ערים לפי קודים' גם אם:
    - יש כותרת בשורה 1
    - שורה ריקה אחרי
    - שורת הכותרות האמיתית רק בשורה 2/3/4...
    מחפש שורה שיש בה גם 'ישוב' וגם 'סמל'.
    """
    df0 = pd.read_excel(city_file, sheet_name=0, header=None)
    header_row = None

    for i, row in df0.iterrows():
        vals = [str(v) for v in row.tolist()]
        joined = " ".join(vals)
        if ("ישוב" in joined) and ("סמל" in joined):
            header_row = i
            break

    if header_row is None:
        raise ValueError("❌ בקובץ הערים לא נמצאה שורת כותרת עם 'ישוב' ו-'סמל ישוב'.")

    headers = [
        "" if pd.isna(v) else str(v).strip()
        for v in df0.iloc[header_row].tolist()
    ]
    df = df0.iloc[header_row + 1:].copy()
    df.columns = headers

    name_col = None
    code_col = None
    for c in df.columns:
        cs = str(c)
        if "ישוב" in cs and "סמל" not in cs and name_col is None:
            name_col = c
        if "סמל" in cs and code_col is None:
            code_col = c

    if name_col is None or code_col is None:
        raise ValueError("❌ בקובץ הערים חייבות להיות עמודות 'ישוב' ו-'סמל ישוב' (או שמות דומים).")

    codes = pd.to_numeric(df[code_col], errors="coerce")
    names = df[name_col].astype(str)

    lut: dict[int, str] = {}
    for code, name in zip(codes, names):
        if pd.isna(code):
            continue
        try:
            lut[int(code)] = name.strip()
        except Exception:
            continue
    return lut


def _load_tree_lut(tree_file) -> dict[int, str]:
    """
    קורא את קובץ 'רשימת עצים לפי קודים' גם אם:
    - יש כותרת גדולה בשורה 1 ('רשימת עצים')
    - יש שורות ריקות באמצע
    - שורת הכותרות האמיתית (Tree, שם עץ וכו') מתחילה רק בשורה 3/4...
    """
    df0 = pd.read_excel(tree_file, sheet_name=0, header=None)

    header_row = None

    # 1) מחפש שורה שיש בה עמודה המכילה 'tree'
    for i, row in df0.iterrows():
        cells = ["" if pd.isna(v) else str(v) for v in row.tolist()]
        if any("tree" in c.lower() for c in cells):
            header_row = i
            break

    # 2) אם לא נמצא — לוקח את השורה הלא־ריקה הראשונה בתור כותרת
    if header_row is None:
        for i, row in df0.iterrows():
            if not pd.isna(row).all():
                header_row = i
                break

    if header_row is None:
        raise ValueError("❌ בקובץ העצים לא נמצאה שורת כותרת מתאימה (עם עמודת Tree).")

    headers = [
        "" if pd.isna(v) else str(v).strip()
        for v in df0.iloc[header_row].tolist()
    ]
    df = df0.iloc[header_row + 1:].copy()
    df.columns = headers

    code_col = None
    name_col = None

    # עמודת הקוד (Tree)
    for c in df.columns:
        if "tree" in str(c).lower():
            code_col = c
            break

    # עמודת שם העץ בעברית (שם עץ)
    for c in df.columns:
        cs = str(c)
        if "שם" in cs and "עץ" in cs:
            name_col = c
            break

    if code_col is None or name_col is None:
        raise ValueError("❌ בקובץ העצים חייבות להיות עמודות 'Tree' ו-'שם עץ' (או שמות דומים).")

    codes = pd.to_numeric(df[code_col], errors="coerce")
    names = df[name_col].astype(str)

    lut: dict[int, str] = {}
    for code, name in zip(codes, names):
        if pd.isna(code):
            continue
        try:
            lut[int(code)] = name.strip()
        except Exception:
            continue
    return lut


def apply_city_tree_lookups(
    merged: pd.DataFrame,
    city_lut: dict[int, str],
    tree_lut: dict[int, str],
) -> pd.DataFrame:
    """
    ממיר קודים → שמות בתוך הדוח הממוזג:
    - בעמודת 'יישוב' (אם יש קודים מספריים)
    - בעמודות מין עץ (אם לפעמים יש שם קוד מספרי במקום שם)
    """
    df = merged.copy()

    # ---- ישובים ----
    if "יישוב" in df.columns:
        codes = pd.to_numeric(df["יישוב"], errors="coerce")
        mask = codes.notna()
        mapped = codes[mask].astype(int).map(city_lut)
        df.loc[mask & mapped.notna(), "יישוב"] = mapped[mapped.notna()]

    # ---- מיני עץ ----
    tree_cols = [
        c for c in df.columns
        if ("שם   מין עץ" in c) or ("שם מין עץ" in c) or (c == "מין עץ")
    ]
    for col in tree_cols:
        codes = pd.to_numeric(df[col], errors="coerce")
        mask = codes.notna()
        mapped = codes[mask].astype(int).map(tree_lut)
        df.loc[mask & mapped.notna(), col] = mapped[mapped.notna()]

    return df


# ---------- סיווג פעולה: כריתה / העתקה ----------

EXCLUDE_PRUNE = r"(דילול|גיזום|תחזוקה|טיפול|חידוש\s*צמרת|עיצוב\s*נוף)"
INCLUDE_CUT = r"(כרית(?:ה|ות)?|כרת|כרות|כריתת)"
INCLUDE_MOVE = r"(העתק(?:ה)?|שימור|transplant|preserv)"


def classify_actions(
    df: pd.DataFrame,
    col_act1: str | None,
    col_act2: str | None,
    col_reason: str | None,
):
    """
    יוצר דגלים בוליאניים:
    - __is_cut__  → שורת כריתה
    - __is_move__ → שורת העתקה/שימור
    """
    pieces = []
    for c in (col_act1, col_act2, col_reason):
        if c and c in df.columns:
            pieces.append(df[c].astype(str))

    if not pieces:
        txt = pd.Series([""] * len(df), index=df.index)
    else:
        txt = pieces[0]
        for p in pieces[1:]:
            txt = txt.str.cat(p, sep=" ", na_rep="")

    txt = (
        txt.str.replace("\u200f", "", regex=False)
           .str.replace("\u200e", "", regex=False)
           .str.strip()
    )

    prune = txt.str.contains(EXCLUDE_PRUNE, case=False, regex=True, na=False)
    cut = txt.str.contains(INCLUDE_CUT, case=False, regex=True, na=False) & ~prune
    move = txt.str.contains(INCLUDE_MOVE, case=False, regex=True, na=False)

    return cut.fillna(False), move.fillna(False)


# ===================== MAIN RUN =====================

if run_btn:
    if not main_file or not city_file or not tree_file:
        st.error("אנא העלה את שלושת הקבצים: דוחות, רשימת ערים, רשימת עצים.")
        st.stop()

    try:
        # 1) טעינת לוקאפ ערים / עצים
        city_lut = _load_city_lut(city_file)
        tree_lut = _load_tree_lut(tree_file)

        # 2) קריאת קובץ הדוחות (מאוחד ללא גליונות מפתח)
        xls = pd.ExcelFile(main_file, engine="openpyxl")

        merged_parts: list[pd.DataFrame] = []
        log_rows: list[dict] = []

        for sname in xls.sheet_names:
            df_raw = pd.read_excel(xls, sheet_name=sname, header=None)
            if df_raw.empty:
                continue

            # כותרת אמיתית של הדוח (חיפוש בשורות הראשונות)
            hdr = detect_header_row(df_raw, scan_rows=8)
            headers = [clean_text(c) for c in df_raw.iloc[hdr].tolist()]
            df = df_raw.iloc[hdr + 1:].reset_index(drop=True)

            if len(headers) < df.shape[1]:
                headers += [f"Unnamed_{i}" for i in range(df.shape[1] - len(headers))]
            headers = headers[: df.shape[1]]
            df.columns = headers

            out = pd.DataFrame(index=df.index, columns=TARGET_COLS)
            used = set()

            for c in df.columns:
                tgt = map_col(c)
                if not tgt:
                    log_rows.append({"sheet": sname, "source_column": c, "mapped_to": ""})
                    continue

                ser = get_series(df, c)

                if tgt == "פעולה":
                    if "פעולה" not in used:
                        out["פעולה"] = ser
                        used.add("פעולה")
                    else:
                        out["פעולה (2)"] = ser
                        used.add("פעולה (2)")
                else:
                    out[tgt] = ser
                    used.add(tgt)

                log_rows.append({"sheet": sname, "source_column": c, "mapped_to": tgt})

            out["__source_sheet__"] = sname
            merged_parts.append(out)

        merged = (
            pd.concat(merged_parts, ignore_index=True)
            if merged_parts else pd.DataFrame(columns=TARGET_COLS)
        )

        # 3) המרת קודים → שמות (יישוב + מין עץ)
        merged = apply_city_tree_lookups(merged, city_lut, tree_lut)

        # 4) פיענוח פעולה/סיבה למלל
        merged = decode_action_reason(merged)

        # 4.5) סידור העמודות כך שהקודים והטקסט יהיו אחד ליד השני
        merged = reorder_columns(merged)

        # 5) תאריכים
        for c in ("מ-תאריך", "עד-תאריך"):
            if c in merged.columns:
                merged[c] = safe_to_datetime_series(merged[c])

        # 6) דגלי כריתה/העתקה
        col_act1 = (
            "פעולה_מפוענחת"
            if "פעולה_מפוענחת" in merged.columns
            else ("פעולה" if "פעולה" in merged.columns else None)
        )
        col_act2 = (
            "פעולה_מפוענחת (2)"
            if "פעולה_מפוענחת (2)" in merged.columns
            else ("פעולה (2)" if "פעולה (2)" in merged.columns else None)
        )
        col_reason = (
            "סיבה_מפוענחת"
            if "סיבה_מפוענחת" in merged.columns
            else (
                "סיבה"
                if "סיבה" in merged.columns
                else ("סיבה  מילולית" if "סיבה  מילולית" in merged.columns else None)
            )
        )

        is_cut, is_move = classify_actions(merged, col_act1, col_act2, col_reason)
        merged["__is_cut__"] = is_cut
        merged["__is_move__"] = is_move

        # 7) כתיבה ל־Excel + עיצוב עמודות
        base = io.BytesIO()
        with pd.ExcelWriter(base, engine="openpyxl") as writer:
            merged.to_excel(writer, sheet_name="Merged", index=False)
            pd.DataFrame({"TargetColumns": TARGET_COLS}).to_excel(
                writer, sheet_name="TargetHeaders", index=False
            )
            pd.DataFrame(log_rows).to_excel(
                writer, sheet_name="MappingLog", index=False
            )

        wb = load_workbook(io.BytesIO(base.getvalue()))
        ws = wb["Merged"]
        header = [c.value for c in ws[1]]

        def _set_col_width(name: str, width: int):
            if name in header:
                idx = header.index(name) + 1
                ws.column_dimensions[get_column_letter(idx)].width = width

        # תאריכים
        for name in ("מ-תאריך", "עד-תאריך"):
            if name in header:
                i = header.index(name) + 1
                for row in ws.iter_rows(min_row=2, min_col=i, max_col=i, max_row=ws.max_row):
                    cell = row[0]
                    cell.number_format = "yyyy-mm-dd"
                    cell.alignment = Alignment(horizontal="right")
                _set_col_width(name, 16)

        # עמודות שימושיות נוספות
        _set_col_width("יישוב", 22)
        _set_col_width("שם   מין עץ", 22)
        _set_col_width("הערות", 26)
        _set_col_width("פעולה_מפוענחת", 14)
        _set_col_width("פעולה_מפוענחת (2)", 16)
        _set_col_width("סיבה_מפוענחת", 16)

        final = io.BytesIO()
        wb.save(final)

        st.success("✅ הקובץ הממוזג והמפוענח מוכן להורדה.")
        st.download_button(
            "⬇️ הורד merged_forest_reports_FINAL_dates_fixed.xlsx",
            data=final.getvalue(),
            file_name="merged_forest_reports_FINAL_dates_fixed.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        with st.expander("תצוגה מקדימה (50 שורות ראשונות)", expanded=False):
            st.dataframe(merged.head(50))

    except Exception as e:
        st.error(f"שגיאה בעיבוד הקבצים: {e}")

else:
    st.info("📎 העלה את שלושת הקבצים ולחץ על הכפתור כדי ליצור קובץ BI מאוחד.")
